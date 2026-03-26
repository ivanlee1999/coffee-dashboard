"""
Excel brew curve parser.

Parses .xlsx exports from coffee scales, normalizes time series data,
detects pour phases, and identifies key brew events.
"""

from io import BytesIO
from datetime import datetime, timedelta

from openpyxl import load_workbook

REQUIRED_COLUMNS = {
    "elapsed",
    "pressure",
    "current_total_shot_weight",
    "flow_in",
    "flow_out",
    "water_temperature_boiler",
    "water_temperature_in",
    "water_temperature_basket",
}

# --- Thresholds for phase detection ---
FLOW_IN_THRESHOLD = 0.3       # ml/s – flow_in above this means water is being poured
MIN_POUR_DURATION_S = 2.0     # seconds – ignore pours shorter than this
MERGE_GAP_S = 3.0             # seconds – merge pours separated by less than this
MIN_SAMPLES_ABOVE = 2         # consecutive samples above threshold to start a pour

# Max plausible brew duration (2 hours). Used to distinguish day-fractions from seconds.
MAX_BREW_SECONDS = 7200.0


def parse_brew_xlsx(file_obj) -> dict:
    """
    Parse an xlsx brew curve file.

    Returns
    -------
    dict with keys:
        points  – list of dicts, each with keys: t, pressure, weight, flow_in,
                  flow_out, temp_boiler, temp_in, temp_basket
        phases  – list of dicts: {name, start_s, end_s}
        events  – list of dicts: {type, time_s}
        series  – pre-built chart series {flow_out: [{t, v}], weight: [{t, v}],
                  flow_in: [{t, v}], pressure: [{t, v}],
                  temp_basket: [{t, v}]}
    """
    rows = _load_rows(file_obj)
    if not rows:
        raise ValueError("Workbook contains no data rows.")
    points = _normalize_rows(rows)
    points = _deduplicate_and_sort(points)
    phases = _detect_pours(points)
    events = _detect_events(points, phases)
    series = _build_series(points)
    return {
        "points": points,
        "phases": phases,
        "events": events,
        "series": series,
    }


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _load_rows(file_obj) -> list[dict]:
    """Read the first worksheet and return a list of row dicts."""
    if isinstance(file_obj, bytes):
        file_obj = BytesIO(file_obj)
    elif hasattr(file_obj, "read"):
        content = file_obj.read()
        file_obj = BytesIO(content)

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().lower() if h else "" for h in header_row]

    # Validate
    found = set(headers)
    missing = REQUIRED_COLUMNS - found
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    col_map = {name: idx for idx, name in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        entry = {}
        for col_name in REQUIRED_COLUMNS:
            entry[col_name] = row[col_map[col_name]]
        rows.append(entry)

    wb.close()
    return rows


def _detect_elapsed_encoding(raw_values: list) -> str:
    """
    Determine how elapsed time is encoded in a batch of raw values.

    Returns one of:
        'seconds'      – values are already in seconds (e.g. 0.071, 5.3, 130.0)
        'day_fraction' – values are Excel day fractions (e.g. 0.000820, 0.001505)
        'string'       – values are strings (handled by _parse_elapsed_string)
        'mixed'        – mixed types, fall back to per-value heuristic
    """
    numerics = [v for v in raw_values if isinstance(v, (int, float)) and v is not None]
    if not numerics:
        return 'string'

    max_val = max(numerics)
    min_val = min(v for v in numerics if v >= 0)

    # If max value is already in a plausible brew-seconds range, treat as seconds
    if max_val >= 1.0:
        return 'seconds'

    # All values are between 0 and 1 — could be day fractions or sub-second intervals
    # Check: does max_val * 86400 fall within a plausible brew time?
    max_as_seconds = max_val * 86400.0
    if 5.0 <= max_as_seconds <= MAX_BREW_SECONDS:
        return 'day_fraction'

    # Very small values that converted are outside plausible range — treat as seconds
    return 'seconds'


def _parse_elapsed_string(s: str) -> float | None:
    """Parse a string elapsed value (HH:MM:SS, MM:SS, or float string)."""
    s = s.strip()
    parts = s.split(":")
    try:
        if len(parts) == 3:
            h, m, sec = parts
            return int(h) * 3600 + int(m) * 60 + float(sec)
        if len(parts) == 2:
            m, sec = parts
            return int(m) * 60 + float(sec)
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_elapsed(value, encoding: str = 'seconds') -> float | None:
    """
    Convert an elapsed value to seconds (float).

    Parameters
    ----------
    value    : raw cell value
    encoding : 'seconds' | 'day_fraction' | 'string' — determined at batch level
    """
    if value is None:
        return None

    if isinstance(value, timedelta):
        return value.total_seconds()

    if isinstance(value, datetime):
        midnight = value.replace(hour=0, minute=0, second=0, microsecond=0)
        return (value - midnight).total_seconds()

    if isinstance(value, (int, float)):
        v = float(value)
        if encoding == 'day_fraction':
            return v * 86400.0
        return v  # 'seconds' or fallback

    # String value
    return _parse_elapsed_string(str(value))


def _safe_float(v, default=0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _normalize_rows(rows: list[dict]) -> list[dict]:
    """Convert raw row dicts to typed point dicts with time in seconds."""
    # Determine encoding from all elapsed values at once (fixes the regression
    # where the old per-value heuristic misclassified day fractions > 0.01).
    raw_elapsed = [r["elapsed"] for r in rows]
    encoding = _detect_elapsed_encoding(raw_elapsed)

    points = []
    for row in rows:
        t = _parse_elapsed(row["elapsed"], encoding=encoding)
        if t is None:
            continue
        points.append({
            "t": round(t, 3),
            "pressure": _safe_float(row["pressure"]),
            "weight": _safe_float(row["current_total_shot_weight"]),
            "flow_in": _safe_float(row["flow_in"]),
            "flow_out": _safe_float(row["flow_out"]),
            "temp_boiler": _safe_float(row["water_temperature_boiler"]),
            "temp_in": _safe_float(row["water_temperature_in"]),
            "temp_basket": _safe_float(row["water_temperature_basket"]),
        })
    return points


def _deduplicate_and_sort(points: list[dict]) -> list[dict]:
    """Sort by time and remove duplicate timestamps (keep last)."""
    points.sort(key=lambda p: p["t"])
    seen: dict[float, int] = {}
    for i, p in enumerate(points):
        seen[p["t"]] = i
    indices = sorted(seen.values())
    return [points[i] for i in indices]


def _smooth(values: list[float], window: int = 5) -> list[float]:
    """Simple moving-average smoothing."""
    if len(values) <= window:
        return values
    out = []
    half = window // 2
    for i in range(len(values)):
        lo = max(0, i - half)
        hi = min(len(values), i + half + 1)
        out.append(sum(values[lo:hi]) / (hi - lo))
    return out


def _detect_pours(points: list[dict]) -> list[dict]:
    """
    Detect pour phases from flow_in signal.

    A pour starts when flow_in exceeds FLOW_IN_THRESHOLD for at least
    MIN_SAMPLES_ABOVE consecutive samples, and ends when it drops below.
    """
    if not points:
        return []

    above = [p["flow_in"] > FLOW_IN_THRESHOLD for p in points]
    regions: list[tuple[int, int]] = []
    in_region = False
    start = 0
    consecutive = 0

    for i, is_above in enumerate(above):
        if is_above:
            if not in_region:
                consecutive += 1
                if consecutive >= MIN_SAMPLES_ABOVE:
                    in_region = True
                    start = i - consecutive + 1
        else:
            if in_region:
                regions.append((start, i - 1))
                in_region = False
            consecutive = 0

    if in_region:
        regions.append((start, len(points) - 1))

    raw_phases = [
        (points[s]["t"], points[e]["t"])
        for s, e in regions
    ]

    raw_phases = [(s, e) for s, e in raw_phases if (e - s) >= MIN_POUR_DURATION_S]

    merged: list[tuple[float, float]] = []
    for s, e in raw_phases:
        if merged and (s - merged[-1][1]) < MERGE_GAP_S:
            merged[-1] = (merged[-1][0], e)
        else:
            merged.append((s, e))

    phases = []
    for i, (s, e) in enumerate(merged):
        name = "Bloom" if i == 0 else f"Pour {i + 1}"
        phases.append({
            "name": name,
            "start_s": round(s, 1),
            "end_s": round(e, 1),
        })

    return phases


def _detect_events(points: list[dict], phases: list[dict]) -> list[dict]:
    """Detect key brew events: bloom end, drawdown points."""
    events = []
    if not phases:
        return events

    events.append({"type": "bloom_end", "time_s": phases[0]["end_s"]})

    for i in range(len(phases) - 1):
        gap_start = phases[i]["end_s"]
        gap_end = phases[i + 1]["start_s"]
        if gap_end - gap_start > 2:
            events.append({
                "type": "drawdown",
                "time_s": round((gap_start + gap_end) / 2, 1),
            })

    for phase in phases:
        events.append({"type": "pour_start", "time_s": phase["start_s"]})
        events.append({"type": "pour_end", "time_s": phase["end_s"]})

    return events


def _build_series(points: list[dict]) -> dict:
    """Build chart-ready series from parsed points."""
    ts = [p["t"] for p in points]
    flow_out_smooth = _smooth([p["flow_out"] for p in points])
    flow_in_smooth = _smooth([p["flow_in"] for p in points])

    return {
        "flow_out": [{"t": t, "v": round(v, 2)} for t, v in zip(ts, flow_out_smooth)],
        "flow_in": [{"t": t, "v": round(v, 2)} for t, v in zip(ts, flow_in_smooth)],
        "weight": [{"t": p["t"], "v": round(p["weight"], 1)} for p in points],
        "pressure": [{"t": p["t"], "v": round(p["pressure"], 2)} for p in points],
        "temp_basket": [{"t": p["t"], "v": round(p["temp_basket"], 1)} for p in points],
    }
